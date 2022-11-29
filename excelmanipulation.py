#  Para poder utilizar openpyxl debemos poner comando pip install openpyxl
#  Ctrl + Shift + p para poder acceder al Virtual Enviroment y poder elegir una libreria
#  /////////////////////// EXCEL ///////////////////////
from openpyxl import Workbook
wb = Workbook()

#  Crear un nuevo archivo Excel
ws = wb.active

#  Cambiar titulo de la hoja
ws.title = "My Sheet"

#  You can review the names of all worksheets of the workbook with the Workbook.sheetname attribute

print(wb.sheetnames)
#['Sheet2', 'New Title', 'Sheet1']

#  You can loop through worksheets
for ws in wb:
  print(ws.title)

#  Asignar valores
ws['A1'] = 42

#  Los registros asignados
ws.append(["Hola"])

#  Asignar tipos de datos Python en Excel
import datetime
ws['A3'] = datetime.datetime.now()

#   *****Control for the new code*****
nums = [4, 78, 9, 84]
for n in nums:
    print(n)

'''Playing with data
Accessing one cell
Now we know how to get a worksheet, we can start modifying cells content. Cells can be accessed directly as keys of the worksheet:'''

#  This will return the cell at A4, or create one if it does not exist yet. Values can be directly assigned:
c = ws['A1']
print(c)

#  This will create the cell at A4, or create one if it does not exist yet. Values can be directly assigned:
ws['A5'] = 4

#  There is also the Worksheet.cell() method.
#  This provides access to cells using row and column notation:

d = ws.cell(row=4, column=2, value=10)
print(d)

#  If I only want the value of a column o row then use:
for row in ws.values:
   for value in row:
     print(value)

#  Add new row
ws.insert_rows(idx=1, amount=5)

ws['A1'] = "clientid"

#  Asignar nuevas hojas al archivo Excel en ws1
ws1 = wb.create_sheet("My New Sheet")
#  Guardar el archivo en cualquier ruta
wb.save("EjemploExcel1.xlsx")

#  ================== Problema a resolver ==================
"""
Asi se puede hacer un comentario multi-linea opcion 1 en Python
"""

'''
Asi se puede hacer un comentario multi-linea opcion 2 en Python
'''

#  Los requerimientos del sistema: Se debe crear un script
'''
1-Primero se debe de acceder a la carpeta donde contendra todos los 100 archivos Excel
2-Acceder a cada uno de los archivos y ahi or medio de un for aplicar la addicion de una fila
3-Este For contendra: addicion de una fila, el nombre de ese valor va a ser "clientid",
  Una vez agregada esa nueva fila va a ser necesario guardar el archivo con el mismo nombre
  O bien con un nuevo nombre y nuevo tipo de archivo csv y no xslx, finalmente cerrar ese
  archivo iterando hasta finlaizar con la cantidad tota de los Excel dentro de la carpeta
'''

#Bueno solo un comentario